import sys
import os.path
from PyQt5 import QtWidgets as qtw
from PyQt5 import QtCore as qtc
from PyQt5 import QtGui as qtg
from PyQt5 import uic,QtGui
import pandas as pd
import numpy as np
from functools import partial
import MagCom_functions as MC
from time import ctime
import matplotlib.pyplot as plt
from matplotlib import cm
import matplotlib.ticker as plticker
from matplotlib.offsetbox import AnchoredText
from openpyxl import load_workbook 
from math import sqrt,atan,tan,asin,floor, pi, sin, cos, degrees, radians, pow, atan2
from clusters_functions import get_distances_matrix,get_time_diff_matrix, get_dst_matrix, get_D, get_if_clustered_matrix, get_clusters, classify_ids_not_specific, year_fraction,calculate_lengths_frequencies

MyWindow, base_class = uic.loadUiType(os.path.join(sys.path[0], 'Complex.ui'))
class MainWindow(base_class):
    def __init__(self, *args,**kwargs):
        super().__init__(*args,**kwargs)
        self.ui = MyWindow()
        self.ui.setupUi(self)
        
        


  
    
        


        
    


class Tab_Magnitude_Completeness():
    def __init__(self, *args,**kwargs):
        
        mw.ui.load_file_button.clicked.connect(self.load_catalog)

        mw.ui.calculate_all_methods_button.clicked.connect(lambda: self.clicked_methods('maxcgftllsemr'))
        mw.ui.calculate_maxc_button.clicked.connect(lambda: self.clicked_methods('maxc'))
        mw.ui.calculate_gft_button.clicked.connect(lambda: self.clicked_methods('gft'))
        mw.ui.calculate_lls_button.clicked.connect(lambda: self.clicked_methods('lls'))
        mw.ui.calculate_emr_button.clicked.connect(lambda: self.clicked_methods('emr'))

        mw.ui.show_graph_button.clicked.connect(self.clicked_show_graph)

        mw.ui.Mag_column_label.hide()
        mw.ui.Mag_column_lineedit.hide()
    def load_catalog(self):
        
        def drop_to_default():
            mw.ui.chosen_sheet_status.setText('Лист не выбран')
            mw.ui.available_sheets_label.setText('Доступные листы:')
            for widget in mw.ui.RightMenu.findChildren((qtw.QPushButton,qtw.QCheckBox)):
                widget.setEnabled(False)
            for label in mw.ui.methods_form.findChildren(qtw.QLabel):
                label.setText('0')

        file_path = qtw.QFileDialog.getOpenFileName(mw, 'Open File',filter="Excel Catalogs (*.xlsx)")
        if not file_path[0]:
            return None
        
        self.filename_to_open = os.path.basename(file_path[0])
        mw.ui.chosen_catalog_label.setText('Выбранный каталог: ' + self.filename_to_open)
        self.chosen_catalog_path = file_path[0]
        self.add_available_sheets()
        drop_to_default()
        
        mw.ui.Mag_column_label.show()
        mw.ui.Mag_column_lineedit.show()
    
    def add_available_sheets(self):
        available_sheets = pd.ExcelFile(self.chosen_catalog_path).sheet_names
        #deleting before adding new sheets
        for i in range(mw.ui.sheets_layout.count()):
            mw.ui.sheets_layout.removeWidget(mw.ui.sheets_layout.itemAt(0).widget())
        #adding sheets
        for sheet in available_sheets:
            mw.ui.sheets_layout.addWidget(qtw.QPushButton(str(sheet)))
        #connecting every sheet button
        items = [mw.ui.sheets_layout.itemAt(i) for i in range(mw.ui.sheets_layout.count())]
        for i,layout_element in enumerate(items):
            widget = layout_element.widget()
            sheet_name = widget.text()
            widget.clicked.connect(partial(self.clicked_chosen_sheet, i,sheet_name))
        
    def clicked_chosen_sheet(self,index,sheet_name):
        
        column_name = mw.ui.Mag_column_lineedit.text()
        try:
            self.mag = MC.simple_read(self.chosen_catalog_path,index, column_name)
        except:
            qtw.QMessageBox.critical(mw,'Ошибка',f'Не найден столбец магнитуд {column_name} на листе {sheet_name}')
            return None
            
        self.chosen_sheet_name = sheet_name
        self.mag_values,self.discrete_counts,self.cumulative_counts = MC.calculate(self.mag)
        for widget in mw.ui.RightMenu.findChildren((qtw.QPushButton,qtw.QCheckBox)):
            widget.setEnabled(True)
        mw.ui.chosen_sheet_status.setText(f'{sheet_name};{column_name}')
        
    def clicked_methods(self,method):
        
        if 'maxc' in method:
            M_MAXC = MC.MAXC(self.mag_values,self.discrete_counts)
            mw.ui.calculate_maxc_label.setText(str(M_MAXC))
        if 'gft' in method:
            M_GFT = MC.Goodness_of_fit(self.mag,self.mag_values,self.discrete_counts,self.cumulative_counts)
            mw.ui.calculate_gft_label.setText(str(M_GFT))
        if 'lls' in method:
            M_LLS = MC.LLS(self.mag,self.mag_values)
            mw.ui.calculate_lls_label.setText(str(M_LLS))
        if 'emr' in method:
            M_EMR = MC.EMR(self.mag,self.mag_values)
            mw.ui.calculate_emr_label.setText(str(M_EMR))
    
    def clicked_show_graph(self):
        
        Mc_results = [wid.text() for wid in mw.ui.methods_form.findChildren(qtw.QLabel)]
        for i,Mc in enumerate(Mc_results):
            try:
                Mc = float(Mc)
            except:
                Mc=0
            Mc_results[i] = Mc
        self.Mc_list = Mc_results
        MC.draw(self.mag_values,self.discrete_counts,self.cumulative_counts,Tab0,mw)




class Tab_orthoregress():
    def __init__(self, *args,**kwargs):
        mw.ui.orth_load_file_button.clicked.connect(self.load_catalog)
        mw.ui.orth_calculate_button.clicked.connect(self.calculate_statistics)
        mw.ui.orth_plot_button.clicked.connect(self.make_plot)

    def load_catalog(self):
        

        file_path = qtw.QFileDialog.getOpenFileName(mw, 'Open File',filter="Excel Catalogs (*.xlsx)")
        if not file_path[0]:
            return None

        self.filename_to_open = os.path.basename(file_path[0])
        mw.ui.orth_chosen_catalog_label.setText('Выбранный каталог: ' + self.filename_to_open)
        self.chosen_catalog_path = file_path[0]

        wb = load_workbook(self.chosen_catalog_path)
        list_name = 'calculate_orthoregress'
        sh = wb[list_name]

        self.x=[]
        self.y=[]  
        for row in sh.iter_rows(2,sh.max_row):
            x_value = row[1].value
            y_value = row[0].value
            if type(x_value) in [int,float] and type(y_value) in [int,float]: 
                self.x.append(row[1].value)
                self.y.append(row[0].value)

        mw.ui.orth_catalog_size_label.setText(f'Размер: {len(self.x)}')
        for widget in mw.ui.orth_rightMenu.findChildren((qtw.QPushButton,qtw.QTextEdit)):
                widget.setEnabled(True)
    def calculate_statistics(self):
        from math import sqrt,atan,tan,asin,floor

        x,y = self.x, self.y
        N = len(x)

        averagex = sum(x)/N #mean - среднее арифм.
        averagey = sum(y)/N #ok
        
        dispx = sum([(x[i]-averagex)**2 for i in range(N)])/N # pvariance - дисперсия
        dispy = sum([(y[i]-averagey)**2 for i in range(N)])/N #ok
        
        self.sigmax = sqrt(dispx)
        self.sigmay = sqrt(dispy) #ok

        self.r=sum([(x[i]-averagex)*(y[i]-averagey) for i in range(N)])/(N*self.sigmax*self.sigmay) #ok correlation - коэф. корреляции Пирсона 
        
        theta = atan((2*self.r*self.sigmax*self.sigmay)/(dispx-dispy))/2
        if dispx<dispy:
            theta+=pi/2

        self.a=tan(theta)
        self.b = averagey-self.a*averagex
        #TODO check if value is correct
        
        numerator = abs(dispx*dispy-self.r**2*self.sigmax*self.sigmay)
        denominator = (N-2)*((dispx-dispy)**2+4*self.r**2*self.sigmax*self.sigmay)
        
        t= float(mw.ui.t_criterion_value_lineedit.text()) # влияет только на погрешность
        self.delta_theta = 1/2*asin(2*t*sqrt(numerator/denominator))

        self.a_lower = tan(theta-self.delta_theta)
        self.a_higher = tan(theta+self.delta_theta)
        self.b_lower = averagey-averagex*self.a_lower
        self.b_higher = averagey - averagex*self.a_higher
        
        self.show_result()
    
    def show_result(self):
        
        mw.ui.orth_result_textEdit.setText(
        "Уравнение линии:\ny= {}*x {:+}\n".format(round(self.a,3),round(self.b,3)) + 
        f"""{round(self.a_lower,4)} < a < {round(self.a_higher,4)}\n{round(self.b_higher,4)} < b < {round(self.b_lower,4)}\nΔθ= {round(self.delta_theta,4)} рад = {str(round(self.delta_theta*180/pi,4))}°\nσ(x)= {round(self.sigmax,3)}\nσ(y)= {round(self.sigmay,3)}\nr^2= {round(self.r**2,3)}"""
        )

    def make_plot(self):
        def get_intercepts(x,y):
    
            N=len(x)
            counters = [1 for i in range(N)]
            checked_pairs=[]
            for i in range(N):
                counter = 0
                ids=[]

                for j in range(N):
                    if y[j]==y[i] and x[j]==x[i] and (x[j],y[j]) not in checked_pairs:
                        counter+=1
                        ids.append(j)
                for id in ids:
                    counters[id]=counter
                checked_pairs.append((x[j],y[j]))
            return counters
        intercepts = get_intercepts(self.x,self.y)

        left = min(self.x)
        right = max(self.x)
        N = len(self.x)
        plt.xlabel('x',fontsize=15)
        plt.ylabel('y',fontsize=15)
        ax = plt.gca()
        points = ax.scatter(self.x,self.y,c=intercepts,cmap=cm.get_cmap('cool',4),marker='o',s=45)
        # plt.plot([0,6],[c,6*m+c],color = 'black') #orthoreg
        ax.plot([left,right],[left*self.a+self.b,right*self.a+self.b],color = 'purple') #article line
    

        # ax.plot([left,right],[left*a_lower+b_higher,right*a_lower+b_higher],color = 'purple',linestyle='dotted') #article lower border
        # ax.plot([left,right],[left*a_higher+b_lower,right*a_higher+b_lower],color = 'purple',linestyle='dotted') #article upper border
        

        uravn_text = AnchoredText(f"y={round(self.a,3)}*x {round(self.b,3):+}\n{N=}", 
                        prop=dict(size=11), frameon=True,loc='upper left',
                        )
        uravn_text.patch.set_boxstyle("round,pad=0.,rounding_size=0.2")
        # if mode:
        #     ax.set_aspect('equal', adjustable='box')
        

        loc = plticker.MultipleLocator(base=0.5) # this locator puts ticks at regular intervals
        # ax.xaxis.set_minor_locator(loc)
        ax.yaxis.set_major_locator(loc)
        ax.add_artist(uravn_text)
        ticks=np.linspace(1,max(intercepts),5,endpoint=True) # установка тиков
        labels = [str(floor(num)) for num in ticks] #округление значение тиков вниз
        bar = plt.colorbar(points, shrink=0.5, aspect=5,ticks=ticks)
        bar.set_ticklabels(labels)
        
        # plt.colorbar.set_yticklabels(['a','b','c','d'])
        plt.grid()
        plt.show()

class Tab_distances(): 
    def __init__(self, *args,**kwargs):
        mw.ui.dist_load_file_button.clicked.connect(self.load_catalog)
        mw.ui.dist_calculate_button.clicked.connect(self.write_to_file)
        
    def load_catalog(self):
        
        from distance_functions import get_distance

        file_path = qtw.QFileDialog.getOpenFileName(mw, 'Open File',filter="Excel Catalogs (*.xlsx)")
        if not file_path[0]:
            return None

        self.filename_to_open = os.path.basename(file_path[0])
        mw.ui.dist_chosen_catalog_label.setText('Выбранный каталог: ' + self.filename_to_open)
        self.chosen_catalog_path = file_path[0]

        wb = load_workbook(self.chosen_catalog_path)
        ws = wb.worksheets[0]
        
        self.distances = ['%.3f\n' %  get_distance(row[1].value, row[2].value, row[3].value, row[4].value) for row in ws.iter_rows(2,ws.max_row)] #very sorry if that's not understandable in the future

        mw.ui.dist_catalog_size_label.setText(f'Пар точек: {len(self.distances)}') # TODO
        for widget in mw.ui.dist_rightMenu.findChildren((qtw.QPushButton)):
            widget.setEnabled(True)
    
    def write_to_file(self):
        
        address = os.path.join(os.getcwd(), 'Distances.txt')
        with open(address, 'a',encoding='utf-8') as file1:
            file1.write(str(ctime()))
            file1.write('\nРасстояния посчитаны в километрах\n')
            for dist in self.distances:
                file1.write(dist)
                
            file1.write("-----------------------------------------\n")
        qtw.QMessageBox.information(mw,'Успех',f'{len(self.distances)} расстояний записано')

class Tab_clusters():
    def __init__(self, *args, **kwargs):
        
        mw.ui.cluster_load_file_button.clicked.connect(self.read_catalog)
        mw.ui.cluster_calculate_button.clicked.connect(self.calculate_clusters)
        mw.ui.cluster_plot_button.clicked.connect(self.plot_clusters)
        mw.ui.cluster_write_to_excel_button.clicked.connect(self.write_to_excel)
        mw.ui.cluster_plot_hist_button.clicked.connect(self.plot_hist)
        # more
    
    def read_catalog(self):

        file_path = qtw.QFileDialog.getOpenFileName(mw, 'Open File',filter="Excel Catalogs (*.xlsx)")
        if not file_path[0]:
            return None

        self.filename_to_open = os.path.basename(file_path[0])
        mw.ui.cluster_chosen_catalog_label.setText('Выбранный каталог: ' + self.filename_to_open)
        self.chosen_catalog_path = file_path[0]
        self.data = pd.read_excel(self.chosen_catalog_path)

        self.latitudes = self.data['Lat']
        self.longitudes = self.data['Lon']
        self.dates =  self.data['Date']
        self.times =  self.data['Origin time']
        self.N = len(self.longitudes)
        
        mw.ui.cluster_catalog_size_label.setText(f'Событий: {self.N}') 

        mw.ui.cluster_calculate_button.setEnabled(True) # после чтения каталога появляется возможность считать кластеры
        

    def calculate_clusters(self):
        
        self.distance_matrix = get_distances_matrix(self.N, self.latitudes, self.longitudes)
        self.datetimes, self.time_difference_matrix = get_time_diff_matrix(self.N, self.dates, self.times)
        self.dst_matrix = get_dst_matrix(self.N, self.distance_matrix, self.time_difference_matrix)
        self.D = get_D(self.dst_matrix)
        self.if_clustered_matrix=  get_if_clustered_matrix(self.N, self.dst_matrix, self.D)
        self.Clusters = get_clusters(self.if_clustered_matrix)
        
        self.ids_of_swarms_exact, self.ids_of_swarms, self.ids_not_swarms = classify_ids_not_specific(self.Clusters)
        self.N_swarms = len(self.ids_of_swarms_exact)

        for widget in mw.ui.cluster_rightMenu.findChildren((qtw.QPushButton, qtw.QTextEdit, qtw.QCheckBox, qtw.QRadioButton )):
            widget.setEnabled(True)

        self.fill_result_textEdit()
        

    def fill_result_textEdit(self):
        Text=f'Найдено {self.N_swarms} роёв\n'
        ML = self.data['ML']
        for i in range(self.N_swarms):
            N_events = len(self.ids_of_swarms_exact[i])
            min_lat = min(self.latitudes[self.ids_of_swarms_exact[i]])
            max_lat = max(self.latitudes[self.ids_of_swarms_exact[i]])
            min_lon = min(self.longitudes[self.ids_of_swarms_exact[i]])
            max_lon = max(self.longitudes[self.ids_of_swarms_exact[i]])
            min_ML = min(ML[self.ids_of_swarms_exact[i]])
            max_ML = max(ML[self.ids_of_swarms_exact[i]])
            Text+=f'{i+1})\nЧисло событий: {N_events}\nШирота: {min_lat}-{max_lat}\nДолгота: {min_lon}-{max_lon}\nML: {min_ML}-{max_ML}\n'
            
        mw.ui.cluster_result_textEdit.setText(Text)


    def plot_clusters(self):
        if mw.ui.cluster_Longitude_time_radioButton.isChecked():
            self.plot_clusters_long_time()
        else:
            self.plot_clusters_long_lat()

    def plot_clusters_long_time(self):
        
        Year_fractions_swarms = [year_fraction(my_datetime) for my_datetime in self.datetimes[self.ids_of_swarms] ]
        Year_fractions_not_swarms = [year_fraction(my_datetime) for my_datetime in self.datetimes[self.ids_not_swarms] ]
        Year_fractions_all = [year_fraction(my_datetime) for my_datetime in self.datetimes]
        fig, ax = plt.subplots()
       

        plt.scatter(self.longitudes[self.ids_not_swarms], Year_fractions_not_swarms, color='blue')
        plt.scatter(self.longitudes[self.ids_of_swarms], Year_fractions_swarms, color='red')

        

        if mw.ui.cluster_if_plot_ids_checkbox.isChecked():
            for i in range(self.N):    # текстовые подписи индексов
                plt.text(self.longitudes[i], Year_fractions_all[i],i,fontsize=10)
        
        plt.grid()
        plt.yticks(np.arange(min(self.dates).year, max(self.dates).year+2,1),size=25)
        plt.xticks(size=25)
        plt.ylabel('Год',size=30)
        plt.xlabel('Долгота, °',size=30)
        ax.ticklabel_format(useOffset=False)
        plt.xlim(-10, 135)
        plt.ylim(min(self.dates).year-1, max(self.dates).year+1.5)
        plt.show()

    def plot_clusters_long_lat(self):
        fig, ax = plt.subplots()
        plt.scatter(self.longitudes[self.ids_not_swarms], self.latitudes[self.ids_not_swarms], color='blue')
        plt.scatter(self.longitudes[self.ids_of_swarms], self.latitudes[self.ids_of_swarms], color='red')   # географическая плоскость

        if mw.ui.cluster_if_plot_ids_checkbox.isChecked():
            for i in range(self.N):    # текстовые подписи индексов
                plt.text(self.longitudes[i], self.latitudes[i],i,fontsize=10)

        plt.ylabel('Широта, °',size=30)
        plt.xlabel('Долгота, °',size=30)
        plt.xticks(size = 25)
        plt.yticks(size = 25)
        plt.grid()
        plt.show()

    def plot_hist(self):
        lengths,bins = calculate_lengths_frequencies(self.Clusters) 
        n, bins, patches = plt.hist(lengths,bins=bins)
        print(n)
        print(bins)
        plt.xlabel('Событий в кластере',size=30)
        plt.ylabel('Количество кластеров',size=30)
        plt.xticks(bins+0.5,size=25) #целые числа начиная с 1
        plt.yticks(n,size=25)  # количества кластеров, set
        for i,p in zip(np.arange(0,max(bins),1),patches):  #цвет слева оранжевый, справа красный
            if i<7:
                plt.setp(p, 'facecolor', 'orange', edgecolor='black')
            else:
                plt.setp(p, 'facecolor', 'red', edgecolor='black')
        plt.show()
    

    
    def write_to_excel(self):
        #TODO отрегулировать ширину полей
        
        N_swarms = len(self.ids_of_swarms_exact)
        with pd.ExcelWriter('Calculated_clusters.xlsx') as writer:
            for i in range(N_swarms):
                df_cur = self.data.iloc[self.ids_of_swarms_exact[i]] # выделяем те строки, которые принадлежат одному рою
                df_cur['Date'] = pd.to_datetime(df_cur['Date']) # переводим в тип датавремя
                df_cur['Date']=df_cur['Date'].dt.strftime('%d.%m.%Y') # чтобы совпадал формат в экселе
                df_cur.drop(df_cur.filter(regex='Unnamed').columns, axis=1, inplace=True) # удаляем лишние столбы без названия
                df_cur.to_excel(writer, sheet_name=f'swarm_{i+1}') # записываем в лист с названием swarm




        
        
if __name__ == '__main__':
    app = qtw.QApplication(sys.argv)
    app.setStyleSheet("""
    QLabel {
    font-size: 20px;
    font-family: consolas;
    }
    QCheckBox {
    font-size: 17px;
    font-family: consolas;  
        }
    QPushButton {
         font-size: 17px;
    font-family: consolas; 
        }""")
    mw = MainWindow()
    Tab0 = Tab_Magnitude_Completeness()
    Tab1 = Tab_orthoregress()
    Tab2 = Tab_distances()
    Tab3 = Tab_clusters()
    mw.show()
    sys.exit(app.exec_())

    #[363, 611, 676, 613, 1189, 1188, 1167, 531, 533, 530, 612, 775, 677]